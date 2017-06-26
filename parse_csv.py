import xml.etree.ElementTree as ET
import os
import glob
import csv
import sys
from standardize import standardize_data

os.chdir(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(os.path.join(os.getcwd(), 'et_xmlfile-1.0.1'))
#et_xmlfile is an external module that is required for openpyxl
import et_xmlfile

os.chdir(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(os.path.join(os.getcwd(), 'jdcal-1.3'))
#jdcal is an external module that is required for openpyxl
import jdcal

os.chdir(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(os.path.join(os.getcwd(), 'openpyxl-2.4.0'))
#openpyxl is an external module that interacts with microsoft excel files
from openpyxl import Workbook, load_workbook


def parse_csv_eu(sheet_name, input_filename, output_filename):
    '''
    sheet_name: the name of the sheet that this function needs to read in Data_Organize.xlsx
    input_filename: name of the input csv file, can be either relative or absolute path.
    output_filename: name of the output csv file, can be either relative or absolute path.
    '''

    #Read the sheet sheet_name in Data_Organize.xlsx, which is used for column reordering. 
    wb = load_workbook(os.path.join(os.getcwd(), 'Data_Organize.xlsx'))
    ws = wb[sheet_name]

    #column_order is a list that keeps the reordering of columns, i.e., [4, 5, 2] means we want the 4th, 5th, 2nd column from the input file. 
    column_order = []
    #column_name is a list of desired names we want for the columns. 
    column_name = []
    i = 2
    #column_name is the 2nd column in Data_Organize.xlsx, starting from 2nd row
    #column_order is the 3rd column in Data_Organize.xlsx, starting from 2nd row
    while ws.cell(row=i, column=2).value != None:
        column_name.append(ws.cell(row=i, column=2).value)
        column_order.append(ws.cell(row=i, column=3).value)
        i += 1

    #Read the input csv file and store the rows in a list. 
    input_data = []
    with open(os.path.join(os.getcwd(), input_filename), newline='') as csvfile:
        reader = csv.reader(csvfile, delimiter=',', quotechar='|')
        for row in reader:
            input_data.append(row)
    csvfile.close()

    #Loop through each row in the input file, and reorder them by the orders given in column_order
    output_data = []
    for row in input_data:
        output_row = []
        for i in range(0, len(column_order)):
            #If column_order is 0, then it means that the desired column is not in the input file, so we add an empty value to the output row. 
            if column_order[i] == 0:
                output_row.append('')
            else:
                output_row.append(row[int(column_order[i]) - 2])
        output_data.append(output_row)
    #Replace the original header to a new header stored in column_name. 
    output_data[0] = column_name

    #Output and save the result in a csv file. 
    with open(os.path.join(os.getcwd(), output_filename), 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(output_data)
    csvfile.close()
            
    return

'''
Example Usage: 
parse_csv_eu('EU_Catalogue', 'EU_Catalogue_Data_Raw.csv', 'EU_Catalogue_Data.csv')
'''



def parse_csv_nasa(sheet_name, input_filename, output_filename):
    '''
    sheet_name: the name of the sheet that this function needs to read in Data_Organize.xlsx
    input_filename: name of the input csv file, can be either relative or absolute path.
    output_filename: name of the output csv file, can be either relative or absolute path.
    '''

    #Read the sheet sheet_name in Data_Organize.xlsx, which is used for column reordering. 
    wb = load_workbook(os.path.join(os.getcwd(), 'Data_Organize.xlsx'))
    ws = wb[sheet_name]

    #column_order is a list that keeps the reordering of columns, i.e., [4, 5, 2] means we want the 4th, 5th, 2nd column from the input file. 
    column_order = []
    #column_name is a list of desired names we want for the columns. 
    column_name = []
    #comment is a list of comments for further operations. 
    comment = []
    i = 2
    #column_name is the 2nd column in Data_Organize.xlsx, starting from 2nd row
    #column_order is the 3rd column in Data_Organize.xlsx, starting from 2nd row
    while ws.cell(row=i, column=2).value != None:
        column_name.append(ws.cell(row=i, column=2).value)
        column_order.append(ws.cell(row=i, column=3).value)
        if ws.cell(row=i, column=4).value is None:
            comment.append('')
        else:
            comment.append(ws.cell(row=i, column=4).value)
        i += 1

    #Read the input csv file and store the rows in a list. 
    input_data = []
    with open(os.path.join(os.getcwd(), input_filename), newline='') as csvfile:
        reader = csv.reader(csvfile, delimiter=',', quotechar='|')
        for row in reader:
            input_data.append(row)
    csvfile.close()

    #Loop through each row in the input file, and reorder them by the orders given in column_order
    output_data = []
    for row in input_data:
        output_row = []
        for i in range(0, len(column_order)):
            #If column_order is 0, then it means that the desired column is not in the input file, so we add an empty value to the output row.
            if column_order[i] == 0:
                output_row.append('')
            else:
                temp_data = row[int(column_order[i]) - 2]
                if 'positive' in comment[i]:
                    temp_data = temp_data.replace('-', '')
                elif 'month' in comment[i]:
                    temp_data = temp_data[0: temp_data.find('-')]    
                output_row.append(temp_data)
        output_data.append(output_row)
    #Replace the original header to a new header stored in column_name. 
    output_data[0] = column_name

    #Output and save the result in a csv file. 
    with open(os.path.join(os.getcwd(), output_filename), 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(output_data)
    csvfile.close()
            
    return

'''
Example Usage: 
parse_csv_nasa('NASA_Catalogue', 'NASA_Catalogue_Data_Raw.csv', 'NASA_Catalogue_Data.csv')
'''


def parse_csv_eu_star(sheet_name, input_filename, output_filename):
    '''
    sheet_name: the name of the sheet that this function needs to read in Data_Organize.xlsx
    input_filename: name of the input csv file, can be either relative or absolute path.
    output_filename: name of the output csv file, can be either relative or absolute path.
    '''

    #Read the sheet sheet_name in Data_Organize.xlsx, which is used for column reordering. 
    wb = load_workbook(os.path.join(os.getcwd(), 'Data_Organize.xlsx'))
    ws = wb[sheet_name]

    #column_order is a list that keeps the reordering of columns, i.e., [4, 5, 2] means we want the 4th, 5th, 2nd column from the input file. 
    column_order = []
    #column_name is a list of desired names we want for the columns. 
    column_name = []
    i = 2
    #column_name is the 2nd column in Data_Organize.xlsx, starting from 2nd row
    #column_order is the 3rd column in Data_Organize.xlsx, starting from 2nd row
    while ws.cell(row=i, column=2).value != None:
        column_name.append(ws.cell(row=i, column=2).value)
        column_order.append(ws.cell(row=i, column=3).value)
        i += 1

    #Read the input csv file and store the rows in a list. 
    input_data = []
    with open(os.path.join(os.getcwd(), input_filename), newline='') as csvfile:
        reader = csv.reader(csvfile, delimiter=',', quotechar='|')
        for row in reader:
            input_data.append(row)
    csvfile.close()

    #Loop through each row in the input file, and reorder them by the orders given in column_order
    star_name_list = []
    output_data = []
    for row in input_data:
        star_name = row[int(column_order[0]) - 2]
        if (len(star_name) == 0) or (star_name in star_name_list):
            continue
        else:
            star_name_list.append(star_name)
        output_row = []
        for i in range(0, len(column_order)):
            #If column_order is 0, then it means that the desired column is not in the input file, so we add an empty value to the output row. 
            if column_order[i] == 0:
                output_row.append('')
            else:
                output_row.append(row[int(column_order[i]) - 2])
        output_data.append(output_row)
    #Replace the original header to a new header stored in column_name. 
    output_data[0] = column_name

    #Output and save the result in a csv file. 
    with open(os.path.join(os.getcwd(), output_filename), 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(output_data)
    csvfile.close()
            
    return

'''
Example Usage: 
parse_csv_eu_star('EU_Catalogue_Star', 'EU_Catalogue_Data_Raw.csv', 'EU_Catalogue_Star_Data.csv')
'''

def parse_csv_nasa_star(sheet_name, input_filename, output_filename):
    '''
    sheet_name: the name of the sheet that this function needs to read in Data_Organize.xlsx
    input_filename: name of the input csv file, can be either relative or absolute path.
    output_filename: name of the output csv file, can be either relative or absolute path.
    '''

    #Read the sheet sheet_name in Data_Organize.xlsx, which is used for column reordering. 
    wb = load_workbook(os.path.join(os.getcwd(), 'Data_Organize.xlsx'))
    ws = wb[sheet_name]

    #column_order is a list that keeps the reordering of columns, i.e., [4, 5, 2] means we want the 4th, 5th, 2nd column from the input file. 
    column_order = []
    #column_name is a list of desired names we want for the columns. 
    column_name = []
    #comment is a list of comments for further operations. 
    comment = []
    i = 2
    #column_name is the 2nd column in Data_Organize.xlsx, starting from 2nd row
    #column_order is the 3rd column in Data_Organize.xlsx, starting from 2nd row
    while ws.cell(row=i, column=2).value != None:
        column_name.append(ws.cell(row=i, column=2).value)
        column_order.append(ws.cell(row=i, column=3).value)
        if ws.cell(row=i, column=4).value is None:
            comment.append('')
        else:
            comment.append(ws.cell(row=i, column=4).value)
        i += 1

    #Read the input csv file and store the rows in a list. 
    input_data = []
    with open(os.path.join(os.getcwd(), input_filename), newline='') as csvfile:
        reader = csv.reader(csvfile, delimiter=',', quotechar='|')
        for row in reader:
            input_data.append(row)
    csvfile.close()

    #Loop through each row in the input file, and reorder them by the orders given in column_order
    star_name_list = []
    output_data = []
    for row in input_data:
        star_name = row[int(column_order[0]) - 2]
        if (len(star_name) == 0) or (star_name in star_name_list):
            continue
        else:
            star_name_list.append(star_name)
        output_row = []
        for i in range(0, len(column_order)):
            #If column_order is 0, then it means that the desired column is not in the input file, so we add an empty value to the output row.
            if column_order[i] == 0:
                output_row.append('')
            else:
                temp_data = row[int(column_order[i]) - 2]
                if 'positive' in comment[i]:
                    temp_data = temp_data.replace('-', '')  
                output_row.append(temp_data)
        output_data.append(output_row)
    #Replace the original header to a new header stored in column_name. 
    output_data[0] = column_name

    #Output and save the result in a csv file. 
    with open(os.path.join(os.getcwd(), output_filename), 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(output_data)
    csvfile.close()
            
    return

'''
Example Usage: 
parse_csv_nasa_star('NASA_Catalogue_Star', 'NASA_Catalogue_Data_Raw.csv', 'NASA_Catalogue_Star_Data.csv')
'''


def parse_csv_nasa_system(sheet_name, input_filename, output_filename):
    '''
    sheet_name: the name of the sheet that this function needs to read in Data_Organize.xlsx
    input_filename: name of the input csv file, can be either relative or absolute path.
    output_filename: name of the output csv file, can be either relative or absolute path.
    '''

    #Read the sheet sheet_name in Data_Organize.xlsx, which is used for column reordering. 
    wb = load_workbook(os.path.join(os.getcwd(), 'Data_Organize.xlsx'))
    ws = wb[sheet_name]

    #column_order is a list that keeps the reordering of columns, i.e., [4, 5, 2] means we want the 4th, 5th, 2nd column from the input file. 
    column_order = []
    #column_name is a list of desired names we want for the columns. 
    column_name = []
    #comment is a list of comments for further operations. 
    comment = []
    i = 2
    #column_name is the 2nd column in Data_Organize.xlsx, starting from 2nd row
    #column_order is the 3rd column in Data_Organize.xlsx, starting from 2nd row
    while ws.cell(row=i, column=2).value != None:
        column_name.append(ws.cell(row=i, column=2).value)
        column_order.append(ws.cell(row=i, column=3).value)
        if ws.cell(row=i, column=4).value is None:
            comment.append('')
        else:
            comment.append(ws.cell(row=i, column=4).value)
        i += 1

    #Read the input csv file and store the rows in a list. 
    input_data = []
    with open(os.path.join(os.getcwd(), input_filename), newline='') as csvfile:
        reader = csv.reader(csvfile, delimiter=',', quotechar='|')
        for row in reader:
            input_data.append(row)
    csvfile.close()

    #Loop through each row in the input file, and reorder them by the orders given in column_order
    star_name_list = []
    output_data = []
    for row in input_data:
        star_name = row[int(column_order[0]) - 2]
        if (len(star_name) == 0) or (star_name in star_name_list):
            continue
        else:
            star_name_list.append(star_name)
        output_row = []
        for i in range(0, len(column_order)):
            #If column_order is 0, then it means that the desired column is not in the input file, so we add an empty value to the output row.
            if column_order[i] == 0:
                output_row.append('')
            else:
                temp_data = row[int(column_order[i]) - 2]
                if 'Replace' in comment[i]:
                    temp_data = temp_data.replace('h', ' ')
                    temp_data = temp_data.replace('d', ' ')
                    temp_data = temp_data.replace('m', ' ')
                    temp_data = temp_data.replace('s', '')
                if i == 1:
                    temp_data = standardize_data(temp_data)
                output_row.append(temp_data)
        output_data.append(output_row)
    #Replace the original header to a new header stored in column_name. 
    output_data[0] = column_name

    #Output and save the result in a csv file. 
    with open(os.path.join(os.getcwd(), output_filename), 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(output_data)
    csvfile.close()
            
    return



'''
Example Usage:
parse_csv_nasa_system('NASA_Catalogue_System', 'NASA_Catalogue_Data_Raw.csv', 'NASA_Catalogue_System_Data.csv')
'''
